from dataclasses import dataclass
from typing import List
from openpyxl import load_workbook
from datetime import date, datetime
import requests
import tempfile
import os
from models import ReportItem

# ----------------------------
# Program number mapping
# ----------------------------

PROGRAM_MAP = {
    "Cab Retreat": "6100",
    "Cab Food/Prizes": "6653",
    "Laundry": "8206",
    "Initiative": "0200",
    "Storage Assistance": "0330",
    "Socioeconomic Inclusivity Fund": "1350",
    "RAs Budget": "6658",
    "Awards": "6333",
    "Spirit": "6137",
    "Sports": "2400",
    "Beer Bike": "6133",
    "Merch": "5080",
    "Community Service": "0331",
    "Permanent Improvements": "6531",
    "Staff Appreciation": "8601",
    "Associates": "6650",
    "Alumni": "6951",
    "Academic Mentors": "1110",
    "PAAs": "6655",
    "Diversity": "6600",
    "Senior Class": "6136",
    "Junior Class": "6601",
    "Sophomore Class": "6602",
    "Freshmen Class": "6134",
    "Off Campus": "6135",
    "Socials": "4600",
    "BGHS": "6603",
    "Chief Justice": "8202",
    "President": "6656",
    "STRIVE": "6604",
    "RHA": "8604",
    "BakerShake": "3150",
    "Baker Black Caucus": "6605",
    "O-Week": "6000",
}


# ----------------------------
# Helpers
# ----------------------------

def export_google_sheet_to_workbook(sheet_url: str):
    """
    Converts a public Google Sheet URL to an xlsx workbook via export.
    """
    spreadsheet_id = sheet_url.split("/d/")[1].split("/")[0]
    export_url = (
        f"https://docs.google.com/spreadsheets/d/"
        f"{spreadsheet_id}/export?format=xlsx"
    )

    response = requests.get(export_url)
    response.raise_for_status()

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.write(response.content)
    tmp.close()

    wb = load_workbook(tmp.name)
    os.unlink(tmp.name)
    return wb


def parse_mmddyyyy(date_str: str) -> datetime:
    return datetime.strptime(date_str, "%m/%d/%Y")

def parse_receipts(cell_value) -> List[str]:
    if not cell_value:
        return []

    # Split on commas, strip whitespace, drop empties
    return [
        part.strip()
        for part in str(cell_value).split(",")
        if part.strip()
    ]

def extract_date(cell_value) -> date:
    if isinstance(cell_value, datetime):
        return cell_value.date()

    if isinstance(cell_value, date):
        return cell_value

    # Always ISO format: YYYY-MM-DD[ ...]
    return datetime.fromisoformat(str(cell_value).split(" ")[0]).date()

def parse_price(cell_value) -> float:
    """
    Parse price from cell value, handling numeric values with or without decimals.
    """
    if not cell_value:
        return 0.0

    # Handle numeric types directly
    if isinstance(cell_value, (int, float)):
        return float(cell_value)

    # Handle string values - strip whitespace and dollar signs
    price_str = str(cell_value).strip().replace("$", "").replace(",", "")

    try:
        return float(price_str)
    except ValueError:
        return 0.0


# ----------------------------
# Core parser
# ----------------------------

def parse_purchases(
    spreadsheet_link: str,
    cardholder_name: str,
    start_date: str
) -> List[ReportItem]:

    wb = export_google_sheet_to_workbook(spreadsheet_link)
    ws = wb["Purchases 2023-2024"]

    start_dt = datetime.strptime(start_date, "%m/%d/%Y").date()
    report_items: List[ReportItem] = []

    # Iterate bottom → top (skip header)
    for row in range(ws.max_row, 1, -1):
        A_timestamp = ws[f"A{row}"].value
        J_pcard = ws[f"J{row}"].value

        if not A_timestamp:
            continue

        # Timestamp is consistently formatted; extract date portion
        row_date = extract_date(A_timestamp)

        # Stop once we're before the start date
        if row_date < start_dt:
            break

        if J_pcard != cardholder_name:
            continue

        D_receipts = ws[f"D{row}"].value
        E_name = ws[f"E{row}"].value
        F_budget = ws[f"F{row}"].value
        G_endowment = ws[f"G{row}"].value
        I_price = ws[f"I{row}"].value
        L_flyer = ws[f"L{row}"].value
        M_items = ws[f"M{row}"].value
        N_event = ws[f"N{row}"].value
        O_vendor = ws[f"O{row}"].value
        P_needs_affidavit = ws[f"P{row}"].value

        description_parts = [
            str(E_name),
            "Baker College",
            row_date.strftime("%m/%d/%Y"),
            str(N_event),
            str(M_items),
            str(F_budget),
        ]

        if F_budget == "Other" and G_endowment:
            description_parts.append(str(G_endowment))

        description = " | ".join(description_parts)
        activity = PROGRAM_MAP.get(F_budget, "")


        report_items.append(
            ReportItem(
                description=description,
                activity=activity,
                date=row_date,
                price=parse_price(I_price),
                vendor=str(O_vendor) if O_vendor else "",
                receipts=parse_receipts(D_receipts),
                flyer=str(L_flyer) if L_flyer else "",
                needsAffidavit=(P_needs_affidavit == "No")
            )
        )

    return report_items
